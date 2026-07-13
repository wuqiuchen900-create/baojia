# -*- coding: utf-8 -*-
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

files = [
    r'C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx',
    r'C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713-2.xlsx',
]

for fp in files:
    wb = openpyxl.load_workbook(fp, data_only=False)
    ws = wb.active
    print(f"\n========== {fp.split(chr(92))[-1]} ==========")

    # 找 八 综合 标题行 与 合计 标题行
    markers = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3, values_only=False):
        r = row[0].row
        c1 = str(row[0].value or '').strip()
        c2 = str(row[1].value or '').strip()
        if c1 == '八' and '综合' in c2 and 'eight' not in markers:
            markers['eight_z'] = r
        if c1 == '九' and '其它' in c2 and 'nine_q' not in markers:
            markers['nine_q'] = r
        if '合计' in c2 and '小计' not in c2 and 'hu' not in markers:
            markers['hu'] = r
    start = markers.get('eight_z', 1)
    end = markers.get('hu', ws.max_row)
    print(f"扫描八综合→合计 (R{start} → R{end})\n")

    print(f"  {'行':4s}  {'H':2s}  {'col1':6s}  {'col2':30s}  {'col3':8s}  {'col4':4s}  {'col6':50s}")
    print(f"  {'-'*4}  {'-'*2}  {'-'*6}  {'-'*30}  {'-'*8}  {'-'*4}  {'-'*50}")
    for r in range(start, end + 1):
        c1 = str(ws.cell(row=r, column=1).value or '').strip()
        c2 = str(ws.cell(row=r, column=2).value or '').strip()
        c3 = str(ws.cell(row=r, column=3).value or '').strip()
        c4 = str(ws.cell(row=r, column=4).value or '').strip()
        c6 = ws.cell(row=r, column=6).value

        # 是否全空：所有9列都不存在 / 全是空白字符串
        all_cols_empty = True
        for col in range(1, 10):
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip() != '':
                all_cols_empty = False
                break

        h_tag = '空' if all_cols_empty else '  '
        f_tag = ''
        if isinstance(c6, str) and c6.startswith('='):
            f_tag = c6[1:50]
        print(f"  R{r:3d}  {h_tag:2s}  {c1[:6]:6s}  {c2[:30]:30s}  {c3[:8]:8s}  {c4[:4]:4s}  {f_tag:50s}")
