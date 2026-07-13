"""扫今天生成的报价单: 把 合计 行 F / H 列 实际公式 + 所有 小计 行坐标 都列出来"""
import openpyxl

fp = r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx"
wb = openpyxl.load_workbook(fp, data_only=False)

for ws in wb.worksheets:
    print(f"\n==== Sheet: {ws.title} | max_row={ws.max_row} max_col={ws.max_column} ====")
    # 第一步: 列出所有 col2 含"小计"行 + 1行 col2 含"合计"
    anchors = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v is None: continue
        sv = str(v).strip()
        if "小计" in sv or "合计" in sv:
            c1 = ws.cell(r, 1).value
            cf = ws.cell(r, 6).value  # F
            ch = ws.cell(r, 8).value  # H
            print(f"R{r:>4} | col1='{str(c1).strip() if c1 else '.'}' | col2='{sv}' | F={cf} | H={ch}")
            anchors.append((r, sv, cf, ch))
