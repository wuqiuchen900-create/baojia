"""
精确扫描两份文件中"八综合"chrome 区域：
- 找到 "八 综合" 标题行
- 列到该 chrome 段内每行 col1-info / col2-name / col6-formula
- 标出任何空白多列行
- 报告 "往前/往后 / 上下边界"
"""
import openpyxl
import sys

FILES = [
    r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx",
    r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713-2.xlsx",
]

for fp in FILES:
    print(f"\n{'='*72}\n文件: {fp}\n{'='*72}")
    try:
        wb = openpyxl.load_workbook(fp, data_only=False)
    except Exception as e:
        print(f"加载失败: {e}")
        continue

    for ws in wb.worksheets:
        print(f"\n--- Sheet: {ws.title} | max_row={ws.max_row} max_col={ws.max_column} ---")
        # 找 chrome 标题行
        chrome_titles = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if v is None:
                continue
            sv = str(v).strip()
            if "八 " in sv and "综合" in sv and "八 加" not in sv:
                chrome_titles.append(("B_col_chrome_header", r, sv))
            elif sv in ("A 电安装(部份改动)", "B 水安装(部份改动)", "C 其他安装"):
                chrome_titles.append(("B_col_segment_title", r, sv))
            elif sv.startswith("九 ") or sv.startswith("十 "):
                chrome_titles.append(("B_col_post_chrome", r, sv))
            elif sv == "其它" or sv == "九 其它":
                chrome_titles.append(("B_col_post_chrome", r, sv))
            elif sv in ("合计", "小计"):
                chrome_titles.append(("B_col_小计合计", r, sv))

        if not chrome_titles:
            print("没找到任何 '八 ' / 子段标题 (可能模板不匹配)")
            continue

        first_r = chrome_titles[0][1]
        last_r = chrome_titles[-1][1]
        print(f"chrome 区域相关行: {first_r}..{last_r}")
        print(f"  标记点共 {len(chrome_titles)} 个:")
        for kind, r, sv in chrome_titles:
            print(f"    [{r:>4}] {kind:24} col2='{sv}'")

        # 输出从 first_r 起到 last_r+10 止的逐行详情
        print(f"\n逐行细节 (R{first_r} ... R{min(last_r+15, ws.max_row)}):")
        print(f"  {'R':>4} | {'col1样式':13} | col2(name)        | col3  | col5 | col6(formula)")
        print(f"  {'-'*4}-+-{'-'*13}-+-{'-'*17}-+-{'-'*5}-+-{'-'*4}-+-{'-'*20}")
        for r in range(first_r, min(last_r + 16, ws.max_row + 1)):
            c1 = ws.cell(r, 1).value or ""
            c2 = ws.cell(r, 2).value or ""
            c3 = ws.cell(r, 3).value
            c5 = ws.cell(r, 5).value
            c6 = ws.cell(r, 6).value
            c1s = str(c1).strip()[:13]
            c2s = str(c2).strip()[:17]
            c3s = "□" if c3 not in (None, "") else " "
            c5s = "□" if c5 not in (None, "") else " "
            c6s = str(c6)[:20] if c6 not in (None, "") else ""
            print(f"  {r:>4} | {c1s:13} | {c2s:17} | {c3s:^5} | {c5s:^4} | {c6s}")

        # 统计 col2-col6 完全空白行
        empty_runs = []
        cur = None
        cur_start = None
        for r in range(first_r, last_r + 1):
            v_all = [ws.cell(r, c).value for c in range(2, 7)]
            if all(v in (None, "") for v in v_all):
                if cur is None:
                    cur_start = r
                cur = r
            else:
                if cur is not None:
                    empty_runs.append((cur_start, cur))
                    cur = None
                    cur_start = None
        if cur is not None:
            empty_runs.append((cur_start, cur))

        print(f"\n空白行区间 (col2..col6 全空白):")
        if empty_runs:
            for s, e in empty_runs:
                print(f"  R{s}..R{e}  ({e-s+1} 行)")
        else:
            print("  无")
