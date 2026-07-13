# -*- coding: utf-8 -*-
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

files = [
    r'C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx',
    r'C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713-2.xlsx',
]

for fp in files:
    wb = openpyxl.load_workbook(fp, data_only=False)
    ws = wb.active
    print(f"\n========== {fp.split(chr(92))[-1]} ==========")
    print(f"行数: {ws.max_row}")

    # find 八综合 marker
    eight_z = None
    nine_q = None
    hu = None  # 合计
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3, values_only=False):
        r = row[0].row
        c1 = str(row[0].value or '').strip()
        c2 = str(row[1].value or '').strip()
        c3 = str(row[2].value or '').strip()
        if c1 == '八' and '综合' in c2:
            eight_z = r
        if c1 == '九' and '其它' in c2:
            nine_q = r
        if '合计' in c2 and '小计' not in c2 and hu is None:
            hu = r
    print(f"八综合 标题行: R{eight_z}")
    print(f"九其它 标题行: R{nine_q}")
    print(f"合计 行: R{hu}")

    # scan 八综合 → 合计 (excluding 合计 itself)
    if eight_z and (hu or nine_q):
        end_r = hu - 1 if hu else (nine_q - 1 if nine_q else ws.max_row)
        print(f"\n--- 八综合 区段 (R{eight_z} → R{end_r}) ---")
        for r in range(eight_z, end_r + 1):
            r1 = ws.cell(row=r, column=1).value
            r2 = ws.cell(row=r, column=2).value
            r6 = ws.cell(row=r, column=6).value
            r8 = ws.cell(row=r, column=8).value
            marker = ''
            if isinstance(r6, str) and r6.startswith('='):
                marker += f'  F=[{r6[:60]}]'
            if isinstance(r8, str) and r8.startswith('='):
                marker += f'  H=[{r8[:60]}]'
            print(f"  R{r:3d} | {str(r1 or '').strip()[:6]:6s} | {str(r2 or '').strip()[:30]:30s}{marker}")
