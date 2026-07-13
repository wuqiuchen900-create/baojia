"""宽松搜索：不再死磕 '八 ' 前缀，而是扫所有 col2 里有'综合/其它/A电/B水'等字样的行 + 全文搜索 col2 含"综合"
然后以这些锚点为参考展开相邻连续空白"""
import openpyxl

FILES = [
    r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx",
    r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713-2.xlsx",
]

ANCHORS = ["综合", "其它", "电安装", "水安装", "其他安装", "安装(部份改动)",
           "小计", "合计", "开荒", "保洁"]

OUT = r"E:\xiangmu\baojia\chrome_dump2.txt"

with open(OUT, "w", encoding="utf-8") as out:
    for fp in FILES:
        out.write(f"\n{'='*80}\n文件: {fp}\n{'='*80}\n")
        wb = openpyxl.load_workbook(fp, data_only=False)
        for ws in wb.worksheets:
            out.write(f"\n--- Sheet: {ws.title} | max_row={ws.max_row} max_col={ws.max_column} ---\n")
            # 用前 9 列作正文（其它列可能是图）
            # 第一步: 全文搜 col2 含任何锚点的行
            hits = []
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, 2).value
                if v is None: continue
                sv = str(v).strip()
                for a in ANCHORS:
                    if a in sv:
                        hits.append((r, sv))
                        break
            out.write(f"  共 {len(hits)} 行 col2 命中锚点:\n")
            for r, sv in hits:
                # 同时把每行的 col1 / col3 列出来看
                c1 = ws.cell(r, 1).value
                c3 = ws.cell(r, 3).value
                c6 = ws.cell(r, 6).value
                label_c3 = ".." if c3 in (None,"") else f"{str(c3).strip()[:8]}"
                label_c6 = ".." if c6 in (None,"") else f"{str(c6).strip()[:13]}"
                out.write(f"    R{r:>4} | col1={str(c1).strip()[:3] if c1 else '.'} | col2=[{sv[:30]}] | col3={label_c3} | col6={label_c6}\n")

            # 第二步: 在 hits 链上找连续空白区间 (col2..col6 全空)
            out.write(f"\n  空白区间 (col2..col6 全空 + 长度 >= 3):\n")
            # 列所有 R1..max_row 的 col2..col6 全空行
            blank_runs = []
            cur_s = None; cur_e = None
            for r in range(1, ws.max_row + 1):
                vs = [ws.cell(r, c).value for c in range(2, 7)]
                if all(v in (None,"") for v in vs):
                    if cur_s is None: cur_s = r
                    cur_e = r
                else:
                    if cur_s is not None and cur_e - cur_s + 1 >= 3:
                        blank_runs.append((cur_s, cur_e))
                    cur_s = cur_e = None
            if cur_s is not None and cur_e - cur_s + 1 >= 3:
                blank_runs.append((cur_s, cur_e))
            for s, e in blank_runs:
                out.write(f"    R{s}..R{e}  ({e-s+1} 行)\n")

            # 第三步: 输出受用户提问的两个空区起点附近 12 行 + 终点附近 5 行
            targets = []
            for s, e in blank_runs:
                if e - s + 1 >= 30:
                    targets.append((s, e))
            out.write(f"\n  大空白区间细节 (>= 30 行): {targets}\n")
            for s, e in targets:
                out.write(f"\n  == 区间 R{s}..R{e} 头尾各 3 行 ==\n")
                hr_lo = max(1, s-3)
                hr_hi = min(ws.max_row, e+3)
                for r in range(hr_lo, hr_hi + 1):
                    cells = [ws.cell(r, c).value for c in range(1, 10)]
                    parts = []
                    for i, v in enumerate(cells):
                        if v in (None, ""):
                            parts.append(".")
                        else:
                            cap = 18 if i == 1 else 6
                            parts.append(str(v).strip()[:cap])
                    out.write(f"    R{r:>4} | {' | '.join(parts)}\n")

import os
print(f"已写入 {OUT}, 大小 {os.path.getsize(OUT)} bytes")
