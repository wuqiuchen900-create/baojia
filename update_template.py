import msoffcrypto, io, zipfile, os, re, shutil

# Decrypt the updated template
src = '模板/地砖.xlsx'
tmp = 'template_tmp.xlsx'

with open(src, 'rb') as f:
    off = msoffcrypto.OfficeFile(f)
    off.load_key(password='1111')
    dec = io.BytesIO()
    off.decrypt(dec)
    raw = dec.getvalue()
print(f'Decrypted: {len(raw)} bytes')

# Clean zip
with zipfile.ZipFile(io.BytesIO(raw), 'r') as zin:
    all_files = zin.namelist()
    wb_xml = zin.read('xl/workbook.xml').decode('utf-8', errors='replace')
    rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8', errors='replace')
    ct_xml = zin.read('[Content_Types].xml').decode('utf-8', errors='replace')

    # Remove definedNames
    wb_clean = re.sub(r'<definedNames[^>]*>.*?</definedNames>', '', wb_xml, flags=re.DOTALL)
    # Remove Wb048gik sheet
    wb_clean = re.sub(r'<sheet[^>]*name="Wb048gik"[^>]*/>', '', wb_clean)

    # Find virus rId
    m = re.search(r'<sheet[^>]*name="Wb048gik"[^>]*r:id="([^"]+)"', wb_xml)
    virus_rid = m.group(1) if m else None
    virus_sheet_files = [f for f in all_files if 'Wb048gik' in f or '048gik' in f.lower()]

    # Remove virus relationship
    rels_clean = rels_xml
    if virus_rid:
        rels_clean = re.sub(rf'<Relationship[^>]*Id="{re.escape(virus_rid)}"[^>]*/>', '', rels_xml)

    # Remove virus content type
    ct_clean = ct_xml
    for vs in virus_sheet_files:
        ct_clean = re.sub(rf'<Override[^>]*PartName="/{re.escape(vs)}"[^>]*/>', '', ct_clean)

    # Write
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in all_files:
            if item in virus_sheet_files:
                print(f'Skipped: {item}')
                continue
            if item == 'xl/workbook.xml':
                zout.writestr(item, wb_clean.encode('utf-8'))
            elif item == 'xl/_rels/workbook.xml.rels':
                zout.writestr(item, rels_clean.encode('utf-8'))
            elif item == '[Content_Types].xml':
                zout.writestr(item, ct_clean.encode('utf-8'))
            else:
                zout.writestr(item, zin.read(item))

# Copy to final destination
dst = 'BaoJiaCAD/template.xlsx'
shutil.copy(tmp, dst)
os.remove(tmp)

print(f'Saved: {dst} ({os.path.getsize(dst)} bytes)')

# Verify
import openpyxl
wb = openpyxl.load_workbook(dst)
ws = wb[wb.sheetnames[0]]
print(f'Sheet: {wb.sheetnames[0]}, Rows: {ws.max_row}')
# Check Row 9 col3 (quantity) - should be empty/0 after user cleared it
qty = ws.cell(9, 3).value
print(f'Row 9 col3 (quantity): {qty}')
wb.close()
print('Done')
