import msoffcrypto, io, zipfile, os, re, shutil, glob

# Find template file
templates = glob.glob('模板/*.xlsx')
dz_files = [f for f in templates if '地砖' in f or 'dizhuan' in f.lower()]
if not dz_files:
    dz_files = [f for f in templates if os.path.getsize(f) < 200000]
    # Use the one with fewer rows
    pass

# Just list what's available
print("Available templates:")
for f in templates:
    print(f"  {f} ({os.path.getsize(f)} bytes)")

# Use the main 地砖 template (first one)
src = dz_files[0] if dz_files else templates[0]
print(f"\nUsing: {src}")

tmp = 'template_tmp.xlsx'

with open(src, 'rb') as f:
    off = msoffcrypto.OfficeFile(f)
    off.load_key(password='1111')
    dec = io.BytesIO()
    off.decrypt(dec)
    raw = dec.getvalue()
print(f'Decrypted: {len(raw)} bytes')

with zipfile.ZipFile(io.BytesIO(raw), 'r') as zin:
    all_files = zin.namelist()
    wb_xml = zin.read('xl/workbook.xml').decode('utf-8', errors='replace')
    rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8', errors='replace')
    ct_xml = zin.read('[Content_Types].xml').decode('utf-8', errors='replace')

    wb_clean = re.sub(r'<definedNames[^>]*>.*?</definedNames>', '', wb_xml, flags=re.DOTALL)
    wb_clean = re.sub(r'<sheet[^>]*name="Wb048gik"[^>]*/>', '', wb_clean)

    m = re.search(r'<sheet[^>]*name="Wb048gik"[^>]*r:id="([^"]+)"', wb_xml)
    virus_rid = m.group(1) if m else None
    virus_files = [f for f in all_files if 'Wb048gik' in f or '048gik' in f.lower()]

    rels_clean = rels_xml
    if virus_rid:
        rels_clean = re.sub(rf'<Relationship[^>]*Id="{re.escape(virus_rid)}"[^>]*/>', '', rels_xml)

    ct_clean = ct_xml
    for vs in virus_files:
        ct_clean = re.sub(rf'<Override[^>]*PartName="/{re.escape(vs)}"[^>]*/>', '', ct_clean)

    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in all_files:
            if item in virus_files:
                print(f'Skipped: {item}')
                continue
            if item == 'xl/workbook.xml':
                zout.writestr(item, wb_clean.encode('utf-8'))
            elif item == 'xl/_rels/workbook.xml.rels':
                zout.writestr(item, rels_clean.encode('utf-8'))
            elif item == '[Content_Types].xml':
                zout.writestr(item, ct_clean.encode('utf-8'))
            else:
                zout.writestr(item, zin.read(item))

dst = 'BaoJiaCAD/template.xlsx'
shutil.copy(tmp, dst)
os.remove(tmp)
print(f'Saved: {dst} ({os.path.getsize(dst)} bytes)')

import openpyxl
wb = openpyxl.load_workbook(dst)
ws = wb[wb.sheetnames[0]]
print(f'Sheet: {wb.sheetnames[0]}, Rows: {ws.max_row}')
qty = ws.cell(9, 3).value
print(f'Row 9 col3 (quantity): {qty}')
wb.close()
print('Done')
