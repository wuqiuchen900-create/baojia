import msoffcrypto, io, zipfile, os, re

def clean_template(src, out_dir):
    name = os.path.basename(src)
    out = os.path.join(out_dir, name)
    print(f'Cleaning: {name}')
    
    with open(src, 'rb') as f:
        off = msoffcrypto.OfficeFile(f)
        off.load_key(password='1111')
        dec = io.BytesIO()
        off.decrypt(dec)
        raw = dec.getvalue()
    
    with zipfile.ZipFile(io.BytesIO(raw), 'r') as zin:
        all_files = zin.namelist()
        
        # Find virus sheet: look for a sheet whose xml content contains Wb048gik
        virus_sheet_files = []
        for item in all_files:
            if 'sheet' in item.lower() and item.endswith('.xml'):
                try:
                    content = zin.read(item).decode('utf-8', errors='replace')
                    if 'Wb048gik' in content or 'XF.Classic' in content:
                        virus_sheet_files.append(item)
                except:
                    pass
        
        if not virus_sheet_files:
            print(f'  No virus sheet found in zip files')
            virus_sheet_files = []
        else:
            print(f'  Virus sheet files: {virus_sheet_files}')
        
        # Read workbook.xml and find virus sheet entry by name
        wb_xml = zin.read('xl/workbook.xml').decode('utf-8', errors='replace')
        rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8', errors='replace')
        ct_xml = zin.read('[Content_Types].xml').decode('utf-8', errors='replace')
        
        # Find virus sheet rId using regex: <sheet ... name="Wb048gik" ... r:id="rIdN" />
        virus_rId = None
        m = re.search(r'<sheet[^>]*name="Wb048gik"[^>]*r:id="([^"]+)"', wb_xml)
        if m:
            virus_rId = m.group(1)
            print(f'  Virus rId: {virus_rId}')
        
        # Find virus file target from relationships using rId
        virus_target = None
        if virus_rId:
            m = re.search(
                rf'<Relationship[^>]*Id="{re.escape(virus_rId)}"[^>]*Target="([^"]+)"',
                rels_xml
            )
            if m:
                virus_target = m.group(1)
                print(f'  Virus target: {virus_target}')
        
        # If not found by rId, try finding by filename
        if not virus_target and virus_sheet_files:
            for vs in virus_sheet_files:
                short = vs.replace('xl/', '').replace('xl\\', '')
                pat = re.escape(short)
                m = re.search(rf'<Relationship[^>]*Target="{pat}"[^>]*/>', rels_xml)
                if m:
                    virus_target = short
                    break
        
        # Remove definedNames
        wb_clean = re.sub(r'<definedNames[^>]*>.*?</definedNames>', '', wb_xml, flags=re.DOTALL)
        print(f'  definedNames removed: {wb_clean != wb_xml}')
        
        # Remove virus sheet entry from workbook.xml
        wb_clean = re.sub(r'<sheet[^>]*name="Wb048gik"[^>]*/>', '', wb_clean)
        
        # Remove virus relationship from workbook.xml.rels
        rels_clean = rels_xml
        if virus_rId:
            rels_clean = re.sub(
                rf'<Relationship[^>]*Id="{re.escape(virus_rId)}"[^>]*/>',
                '', rels_xml
            )
        elif virus_target:
            rels_clean = re.sub(
                rf'<Relationship[^>]*Target="{re.escape(virus_target)}"[^>]*/>',
                '', rels_xml
            )
        
        # Remove content type override for all virus sheet files
        ct_clean = ct_xml
        for vs in virus_sheet_files:
            part = '/' + vs
            ct_clean = re.sub(
                rf'<Override[^>]*PartName="{re.escape(part)}"[^>]*/>',
                '', ct_clean
            )
        
        # Remove other virus files (like macros)
        virus_extra = [f for f in all_files if 'Wb048gik' in f or 
                       any(kw in f.lower() for kw in ['macro', 'vba', '048gik'])]
        
        # Write clean zip
        with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in all_files:
                if item in virus_sheet_files or item in virus_extra:
                    print(f'  Skipped virus: {item}')
                    continue
                if item == 'xl/workbook.xml':
                    zout.writestr(item, wb_clean.encode('utf-8'))
                elif item == 'xl/_rels/workbook.xml.rels':
                    zout.writestr(item, rels_clean.encode('utf-8'))
                elif item == '[Content_Types].xml':
                    zout.writestr(item, ct_clean.encode('utf-8'))
                else:
                    zout.writestr(item, zin.read(item))
    
    print(f'  Saved: {os.path.getsize(out)} bytes')
    return out

files = [
    '模板/2025年最新巴中帝豪装饰最新报价 - 副本.xlsx',
    '模板/南江半山爱马仕10-101.xlsx',
    '模板/地砖.xlsx',
    '模板/木地板.xlsx',
]

out_dir = '模板_clean'
os.makedirs(out_dir, exist_ok=True)

for f in files:
    result = None
    try:
        result = clean_template(f, out_dir)
    except Exception as e:
        print(f'  ERROR: {e}')

print('\n=== Clean verification ===')
for fname in os.listdir(out_dir):
    if not fname.endswith('.xlsx'): continue
    path = os.path.join(out_dir, fname)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        # Check no virus content
        has_virus = False
        for sn in wb.sheetnames:
            if 'Wb048gik' in sn or '048gik' in sn.lower():
                has_virus = True
        print(f'{fname}: sheets={wb.sheetnames}, rows={ws.max_row}, virus={has_virus}')
        wb.close()
    except Exception as e:
        print(f'{fname}: ERROR {e}')
