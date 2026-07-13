"""
完整比对 两份文件 + 模拟 FixTotalFormula 逻辑:
1. 扫 template.xlsx: 找所有 col2 含"小计"/"合计"行
2. 模拟我的代码逻辑生成 应该的公式 (chunked SUM)
3. 扫 output.xlsx 实际公式
4. 对比差异
"""
import openpyxl

TEMPLATE = r"E:\xiangmu\baojia\BaoJiaCAD\template.xlsx"
OUTPUT = r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx"
CHUNK = 150


def scan(fp, label):
    print(f"\n========= {label}: {fp} =========")
    wb = openpyxl.load_workbook(fp, data_only=False)
    for ws in wb.worksheets:
        print(f"\n--- Sheet: {ws.title}  max_row={ws.max_row} max_col={ws.max_column} ---")

        # 1. 列出 col2 含小计 / 合计 行
        anchors = []
        xj_rows = []
        hj_rows = []
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if v is None: continue
            sv = str(v).strip()
            if "小计" in sv:
                xj_rows.append(r)
                anchors.append((r, "小计", sv))
            if "合计" in sv:
                hj_rows.append(r)
                anchors.append((r, "合计", sv))

        print(f"\n小计行 ({len(xj_rows)} 个): {xj_rows}")
        print(f"合计行 ({len(hj_rows)} 个): {hj_rows}")
        print(f"\n详细:")
        for r, tag, sv in anchors:
            c1 = ws.cell(r, 1).value
            cf = ws.cell(r, 6).value
            ch = ws.cell(r, 8).value
            print(f"  R{r:>4} [{tag}] col1='{str(c1).strip() if c1 else '.'}' | col2='{sv}' | F={cf} | H={ch}")

        # 2. 模拟 FixTotalFormula 应该输出的公式
        if hj_rows:
            total_row = hj_rows[0]  # 第一个合计
            subtotal_rows = xj_rows  # 所有小计行
            print(f"\n模拟 FixTotalFormula 应该输出 (totalRow=R{total_row}, subtotalRows.Count={len(subtotal_rows)}):")

            # 分块
            f_chunks = []
            h_chunks = []
            for i in range(0, len(subtotal_rows), CHUNK):
                chunk = subtotal_rows[i:i+CHUNK]
                f_chunks.append("SUM(" + ",".join(f"F{r}" for r in chunk) + ")")
                h_chunks.append("SUM(" + ",".join(f"H{r}" for r in chunk) + ")")

            final_F = "=" + "+".join(f_chunks)
            final_H = "=" + "+".join(h_chunks)
            print(f"  F 预期: {final_F[:200]}{'...' if len(final_F)>200 else ''}")
            print(f"  H 预期: {final_H[:200]}{'...' if len(final_H)>200 else ''}")

        # 3. 显示 Sheet: this is 第一张 还是 其他
        # 仅第一张 sheet 接续


scan(TEMPLATE, "源 TEMPLATE")
scan(OUTPUT, "输出 报价单")
