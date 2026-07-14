"""
🔧 v6 报价单 验证助手 — 抽取 xlsx 中每层每个 roomType 的 「铺地砖」/「正铺」/「正贴」/「菱铺」/「菱贴」行,
    打印 (楼层, roomType, C2行名, C3数量, C5材质, C7人工), 然后按预期 spec 标红/标绿.

预期场景 (与v6 sim 一致):
  1F 客餐厅: sp750-1500  →  C2=正铺地砖(750*1500MM)  C5=28  C7=71
  2F 客餐厅: NONE        →  C2 保持原模板  C5/C7 保持原模板
  3F 客餐厅: sp600-1200  →  C2=正铺地砖(600*1200MM)  C5=28  C7=48
  4F 客餐厅: NONE        →  C2 保持原模板  C5/C7 保持原模板

使用:  python scripts/verify_xlsx_v6.py "C:\path\to\output.xlsx"
"""
import sys
import os

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    print("ERR: openpyxl 未装.  装: pip install openpyxl")
    sys.exit(1)


# ─────────────────────────────────────────────────────────────────
# 预期场景 (与 v6 sim 一致; 视 实际 config.TileSpecOptions 可调)
# ─────────────────────────────────────────────────────────────────
EXPECTED = {
    # (floor_label, room_type) → list of expected tileish rows
    ("一楼", "客餐厅"): [
        {"c2": "正铺地砖（750*1500MM）", "c5": 28.0, "c7": 71.0, "spec": "sp750-1500"},
    ],
    ("二楼", "客餐厅"): [
        # NONE → 不验证具体值, 只要求 任何 tileish 行的 C2 不 包含 「正铺地砖（750*1500MM）」 这类 override
        {"c2": "原模板行名 (无 override)", "c5": None, "c7": None, "spec": "<NONE>"},
    ],
    ("三楼", "客餐厅"): [
        {"c2": "正铺地砖（600*1200MM）", "c5": 28.0, "c7": 48.0, "spec": "sp600-1200-LR"},
    ],
    ("四楼", "客餐厅"): [
        {"c2": "原模板行名 (无 override)", "c5": None, "c7": None, "spec": "<NONE>"},
    ],
}

FLOOR_LABELS = {"一楼", "二楼", "三楼", "四楼", "五楼", "一层", "二层", "三层", "四层", "五层", "首层"}
TILEISH_KEYWORDS = ("铺地砖", "正铺", "正贴", "菱铺", "菱贴")
ROOM_HEADER_PATTERN = ("客餐厅", "主卧", "卧室", "厨房", "卫生间", "阳台", "外花园")


# ─────────────────────────────────────────────────────────────────
# 扫描工作表 — 按 floorHeader / categoryHeader 分组
# ─────────────────────────────────────────────────────────────────
def scan_sheet(ws):
    """
    Return list of blocks:
        [
            {
                'floor': '一楼',     # 或 '未指定'
                'roomType': '客餐厅',
                'header_row': 5,
                'items': [
                    {'c2': '正铺地砖（...）', 'c3': 18.0, 'c5': 28.0, 'c7': 71.0, 'row': 8},
                    ...
                ],
            },
            ...
        ]
    """
    blocks = []
    current_floor = ""
    current_room = ""
    current_items = []

    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c1s = str(c1 or "").strip()
        c2s = str(c2 or "").strip()

        # floorHeader row (merged 楼层层头)
        if c1s in FLOOR_LABELS and not c2s:
            # commit previous block
            if current_room:
                blocks.append({
                    "floor": current_floor or "未指定",
                    "roomType": current_room,
                    "header_row": r,
                    "items": current_items,
                })
            current_floor = c1s
            current_room = ""
            current_items = []
            continue

        # categoryHeader row (eg "一、客餐厅" with C1="一、", C2="客餐厅"; 或 "客餐厅" 单列头 with C1 空)
        # 🔧 HIGH-fix: 必须 精准匹配, 不允许 c2s.endswith(rt) 的「item name 里包含 roomType」误类比 — 例
        #   "客餐厅及外花园" 是 item row, 不应被 当作 分类表头. 原检测 “c2s.endswith(rt) and len<12” 会 误报.
        is_zh_num_prefix = c1s in "一二三四五六七八九十、"
        if c2s in ROOM_HEADER_PATTERN and (c1s == "" or is_zh_num_prefix):
            if current_room:
                blocks.append({
                    "floor": current_floor or "未指定",
                    "roomType": current_room,
                    "header_row": r,
                    "items": current_items,
                })
            current_room = c2s
            current_items = []
            continue

        # tileish item row
        if any(k in c2s for k in TILEISH_KEYWORDS):
            c3 = ws.cell(r, 3).value
            c5 = ws.cell(r, 5).value
            c7 = ws.cell(r, 7).value
            current_items.append({
                "c2": c2s,
                "c3": c3,
                "c5": c5,
                "c7": c7,
                "row": r,
            })

    # commit last
    if current_room:
        blocks.append({
            "floor": current_floor or "未指定",
            "roomType": current_room,
            "header_row": 0,
            "items": current_items,
        })

    return blocks


# ─────────────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print('用法: python scripts/verify_xlsx_v6.py "C:\\path\\to\\output.xlsx"')
        print('       (拖拽 xlsx 到终端 也行)')
        sys.exit(1)
    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"ERR: 文件不存在: {path}")
        sys.exit(1)
    print(f"📄 读取: {path}")
    print(f"   size: {os.path.getsize(path)} bytes")
    print()

    wb = openpyxl.load_workbook(path, data_only=False)
    print(f"   sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]  # 主工作表

    blocks = scan_sheet(ws)
    print(f"\n🔍 扫到 {len(blocks)} 个 (floor, roomType) 块, {sum(len(b['items']) for b in blocks)} 个 tileish 行\n")

    if not blocks:
        print("⚠ 未扫到任何块 — 可能 worksheet 结构 与预期不同, 手工核对 xlsx.")
        return

    pass_count = 0
    fail_count = 0
    unexpected_count = 0

    for blk in blocks:
        fl, rt = blk["floor"], blk["roomType"]
        for it in blk["items"]:
            tag = f"  [{fl}|{rt}] R{it['row']}"
            print(f"{tag}  C2={it['c2']!r}")
            print(f"         C3={it['c3']}  C5={it['c5']}  C7={it['c7']}")

            exp = EXPECTED.get((fl, rt))
            if exp is None:
                # 未知组合 — 不评判, 只 dump
                unexpected_count += 1
                print(f"         (unexpected (floor, roomType), 跳过预期对账)")
                continue

            # match against expected list
            ok_match = False
            for e in exp:
                if e["c2"] == "原模板行名 (无 override)":
                    # NONE: 只要求 C2 != 任意 override spec 形态
                    if any(s in it["c2"] for s in [
                        "正铺地砖（750*1500MM）", "正铺地砖（600*1200MM）",
                        "正铺地砖（800*1600MM）", "正铺地砖（900*1800MM）",
                        "正铺地砖（300-800MM）", "菱铺", "菱贴",
                    ]):
                        continue
                    if e["c5"] is None and e["c7"] is None:
                        ok_match = True
                    else:
                        ok_match = (it["c5"] == e["c5"] and it["c7"] == e["c7"])
                    if ok_match:
                        break
                else:
                    if it["c2"] == e["c2"] and (
                        e["c5"] is None or float(it["c5"] or 0) == float(e["c5"])
                    ) and (
                        e["c7"] is None or float(it["c7"] or 0) == float(e["c7"])
                    ):
                        ok_match = True
                        break

            if ok_match:
                pass_count += 1
                print(f"         ✅ PASS  (spec={e['spec']})")
            else:
                fail_count += 1
                print(f"         ❌ FAIL  (expected one of: {[e['spec'] for e in exp]})")
            print()

    print(f"\n{'='*60}")
    print(f"📊 结果汇总: PASS={pass_count}  FAIL={fail_count}  unexpected={unexpected_count}")
    print(f"   总 (floor, roomType) 块: {len(blocks)}")
    if fail_count == 0 and unexpected_count == 0:
        print("🎉 全部对账通过! v6 多楼层 spec 选择 已正确落地.")
    elif fail_count == 0:
        print("🟡 实际对账都过了, 但 有 额外 (floor, roomType) 块 — 检查是否 期望在 EXPECTED 表中")
    else:
        print("🔴 有 FAIL — 检查 xlsx 实际 输出 vs 上方 EXPECTED (可能 模板选择/模板结构 假设 不符).")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
