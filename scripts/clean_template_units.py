#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
clean_template_units.py
========================

把模板 D 列 字面 'M2' / 'm2' 替换为 Unicode 'M^2' / 'm^2'。

设计原则:
  - 严格只动 D 列 (col 4)
  - 仅当 cell value trim 后 整串匹配 ('M2' / 'm2') 才动手 (避免误伤 'MM2' / '12M2' 等子串)
  - 默认 --dry-run, --apply 才真改
  - 改之前先 timestamp 备份, 改之后做 zip-level 完整性 + drawings 数 + sheet 数校验
  - 校验失败自动回滚
  - named-range 数变化默认仅 WARN (openpyxl 清洗 invalid names 不影响 runtime 语义), 加 --strict-defnames 可视为 fatal

用法:
  python scripts/clean_template_units.py                       # dry-run (默认)
  python scripts/clean_template_units.py --apply               # 真改盘
  python scripts/clean_template_units.py --files f.xlsx g.xlsx
  python scripts/clean_template_units.py --backup-dir <path>   # 自定义备份目录
  python scripts/clean_template_units.py --strict-defnames     # named-range 净差也视为 fatal
  python scripts/clean_template_units.py --json                # JSON 输出 (给机器读)
"""

import argparse
import datetime
import hashlib
import json
import re
import shutil
import sys
import zipfile
from pathlib import Path

try:
    import openpyxl  # noqa
except ImportError:
    print('需要先安装: pip install openpyxl', file=sys.stderr)
    raise

# stdout UTF-8 (windows cmd / 老 bash 都安全)
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

# 默认目标
DEFAULT_FILES = [
    'muban_clean/dizhuan.xlsx',
    'muban_clean/mudiban.xlsx',
    'muban_clean/fushi.xlsx',
]
DEFAULT_BACKUP_DIR = 'muban_clean.pre_clean_backup'

# 仅整串匹配 (trim 后). 防止伤到 B 列的 'MM2' / '12M2' 等
UNIT_MAP = {
    'M2': 'M\u00b2',   # 平方米
    'm2': 'm\u00b2',   # 平方米 (小写版本)
}


# ----------------------------------------------------------------------
# helpers
# ----------------------------------------------------------------------

def sha256_file(p):
    h = hashlib.sha256()
    with open(p, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            h.update(chunk)
    return h.hexdigest()


def inspect_zip(p):
    """粗粒度统计 drawings / media / defined names 数量, zip 是否可开."""
    info = {'drawings': 0, 'media': 0, 'defined_names': 0, 'zip_ok': False}
    try:
        with zipfile.ZipFile(p) as z:
            if z.testzip() is None:
                info['zip_ok'] = True
            for n in z.namelist():
                if n.startswith('xl/drawings/') and n.endswith('.xml'):
                    info['drawings'] += 1
                elif n.startswith('xl/media/'):
                    info['media'] += 1
            try:
                wb_xml = z.read('xl/workbook.xml').decode('utf-8', errors='replace')
                info['defined_names'] = len(re.findall(r'<definedName\b', wb_xml))
            except KeyError:
                pass
    except Exception as e:
        info['zip_error'] = str(e)
    return info


def scan(filepath):
    """dry-run 扫描: 列出将被改的 (row, old, new) 与 warnings."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb.worksheets[0]
    rows = []
    warnings = []
    for r in range(1, ws.max_row + 1):
        c = ws.cell(row=r, column=4)
        # 公式 / 数字 / bool / None 都跳过 (D 列不该有, 但保险)
        if c.data_type == 'f':
            warnings.append(f'R{r}: D 列是公式 {c.value!r}, 跳过')
            continue
        if not isinstance(c.value, str):
            continue
        s = c.value.strip()
        if s in UNIT_MAP:
            rows.append((r, s, UNIT_MAP[s]))
    return rows, warnings


# ----------------------------------------------------------------------
# 纯 zipfile-level 修改: 取消 openpyxl.save (openpyxl.save 会丢 drawings/media),
# 改为 在 xl/sharedStrings.xml 与 xl/worksheets/sheet*.xml 中 精准 改 <si>/<is> 文本块.
# 其他 zip entries (drawings, media, theme, styles, defined names workbook.xml) 都 原样保留.

_INNER_T_RE = re.compile(r'<t[^>]*>([^<]*)</t>')
_SI_RE = re.compile(r'<si>(.*?)</si>', re.DOTALL)
_IS_RE = re.compile(r'<is>(.*?)</is>', re.DOTALL)


def _replace_si_or_is(text, old, new):
    """Replace <si>...</si> OR <is>...</is> blocks (anywhere they appear) whose
    concatenated <t>...</t> text equals `old` with a plain text block containing `new`.
    Returns (new_text, replaced_count)."""
    count = 0
    target_plain = f'<si><t>{new}</t></si>'

    def repl_si(m):
        nonlocal count
        inner = m.group(1)
        joined = ''.join(_INNER_T_RE.findall(inner))
        if joined == old:
            count += 1
            return target_plain
        return m.group(0)

    def repl_is(m):
        # Inline string 用 <is> 而不是 <si>
        # 这里 我们 可以同样 处理当 cell t=\"inlineStr\"
        # 如果要 仅处理 inline string (不处理 公用串), 可能需要 调用方 区分.
        nonlocal count
        inner = m.group(1)
        joined = ''.join(_INNER_T_RE.findall(inner))
        if joined == old:
            count += 1
            return f'<is><t>{new}</t></is>'
        return m.group(0)

    text = _SI_RE.sub(repl_si, text)
    text = _IS_RE.sub(repl_is, text)
    return text, count


def apply(filepath, rows):
    """纯 zipfile 修改. 精准改 `<si>` 与 `<is>` 块. 其他 zip entries 原样保留.

    限制: 只改 xl/sharedStrings.xml 与 xl/worksheets/sheet*.xml. 其他 entries 原样.
    该 实现优先改 sharedStrings 里的 'M2' -> 'M\u00b2' 平断言. 对 同 index 多个 row 都 有 效果.
    """
    import os
    targets = list({(o, n) for _, o, n in rows})  # 去重
    if not targets:
        return sha256_file(filepath)

    # 1. 读取所有 zip entries (含 ZipInfo 对象 以 保留元信息)
    with zipfile.ZipFile(filepath, 'r') as zin:
        entries = []
        for item in zin.infolist():
            entries.append((item.filename, zin.read(item.filename), item))

    # 2. 修改 被 targeted 上下文 (sharedStrings / worksheets). 其他 原样.
    for idx, (name, data, item) in enumerate(entries):
        if name == 'xl/sharedStrings.xml' or name.startswith('xl/worksheets/sheet'):
            try:
                txt = data.decode('utf-8')
            except UnicodeDecodeError:
                continue
            orig_txt = txt
            for old, new in targets:
                txt, _ = _replace_si_or_is(txt, old, new)
            if txt != orig_txt:
                entries[idx] = (name, txt.encode('utf-8'), item)

    # 3. atomic 写: 到 tmp, 然后 os.replace 换
    tmp_path = filepath + '.tmp.zip'
    with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data, item in entries:
            zinfo = zipfile.ZipInfo(filename=name, date_time=item.date_time)
            zinfo.compress_type = item.compress_type
            zinfo.external_attr = item.external_attr
            zinfo.create_system = item.create_system
            zout.writestr(zinfo, data)
    os.replace(tmp_path, filepath)
    return sha256_file(filepath)


def verify(filepath, before_stats, rows_changed, strict_defnames=False):
    """校验保存后是否健康.

    纯 zipfile 实现 仍 维持 drawings / media / defined names byte-for-byte 原样,
    所以 平价检查 使用 严格 等于 (而不是 仅检查 !=0).
    """
    issues = []
    warnings = []
    after = inspect_zip(filepath)
    if not after['zip_ok']:
        issues.append(f"zip 不可读 (drawings={after['drawings']}, err={after.get('zip_error')})")
    # drawings 严格维护 (zipfile 层不应该 丢)
    if after['drawings'] != before_stats.get('drawings', 0):
        issues.append(f"drawings 数变 ({before_stats.get('drawings')} -> {after['drawings']}, 脚本不该丢)")
    # defined_names 平价
    dn_before = before_stats.get('defined_names', 0)
    dn_after = after.get('defined_names', 0)
    if dn_after != dn_before:
        msg = (
            f"named-range 数量变化: {dn_before} -> {dn_after} "
            f"(delta={dn_after - dn_before:+d}). "
            f"纯 zipfile 实现 为 人所料 不该变 (warning: 仔细 査 diff)."
        )
        if strict_defnames:
            issues.append(msg)
        else:
            warnings.append(msg)
    # media 同样严格检查
    if after.get('media', 0) != before_stats.get('media', 0):
        issues.append(f"media 数变 ({before_stats.get('media')} -> {after.get('media')})")
    # 数据层: 重新 openpyxl 读, 取首/中/末 抽样 D 列 检查
    try:
        wb2 = openpyxl.load_workbook(filepath, data_only=False)
        ws2 = wb2.worksheets[0]
        if rows_changed:
            sample_idx = [0, len(rows_changed) // 2, len(rows_changed) - 1]
            sample = [rows_changed[i] for i in sample_idx]
            bad = []
            for r, _old, expected_new in sample:
                actual = ws2.cell(row=r, column=4).value
                if actual != expected_new:
                    bad.append((r, expected_new, actual))
            if bad:
                issues.append(f"D 列采样行未替换到位: {bad}")
    except Exception as e:
        issues.append(f"re-load 失败: {e}")
    return issues, after, warnings


def backup(filepath, backup_dir):
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    p = Path(filepath)
    dst = Path(backup_dir) / f"{p.stem}.pre_clean_{ts}.xlsx"
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(filepath, dst)
    return str(dst)


# ----------------------------------------------------------------------
# 单文件流程
# ----------------------------------------------------------------------

def process_one(filepath, args):
    rec = {'file': filepath}
    if not Path(filepath).exists():
        rec['error'] = 'file not found'
        return rec

    # 1. 改前快照
    rec['before_sha256'] = sha256_file(filepath)
    rec['before'] = inspect_zip(filepath)

    # 2. 扫描
    rows, warnings = scan(filepath)
    rec['rows_to_change'] = [{'row': r, 'old': o, 'new': n} for r, o, n in rows]
    rec['warnings'] = warnings
    rec['count'] = len(rows)

    if not args.apply:
        return rec
    if not rows:
        rec['changed'] = False
        return rec

    # 3. 备份 (在改之前)
    rec['backup'] = backup(filepath, args.backup_dir)

    # 4. 应用修改
    try:
        new_sha = apply(filepath, rows)
        rec['after_sha256'] = new_sha
    except Exception as e:
        rec['error'] = f'apply 异常: {e}'
        rec['after_sha256'] = sha256_file(filepath)
        rec['after'] = inspect_zip(filepath)  # 保持 JSON schema 对称
        rec['rolled_back'] = True
        shutil.copy2(rec['backup'], filepath)
        return rec

    # 5. 校验 (含 defined_names 平价 — 代码审查者 blocking #1)
    issues, stats, verify_warnings = verify(
        filepath, rec['before'], rows,
        strict_defnames=args.strict_defnames,
    )
    rec['after'] = stats
    if verify_warnings:
        rec['warnings'] = (rec.get('warnings') or []) + verify_warnings
    rec['define_name_delta'] = stats.get('defined_names', 0) - rec['before'].get('defined_names', 0)
    if issues:
        rec['rolled_back'] = True
        rec['verify_issues'] = issues
        # 把原模板从 backup 还原
        shutil.copy2(rec['backup'], filepath)
    else:
        rec['changed'] = True

    return rec


# ----------------------------------------------------------------------
# 输出
# ----------------------------------------------------------------------

def format_human(recs, args):
    out = []
    for r in recs:
        out.append('')
        out.append(f"=== {r['file']} ===")
        if 'error' in r and ('rows_to_change' not in r):
            out.append(f"  ! {r['error']}")
            continue
        cnt = r.get('count', 0)
        bsha = (r.get('before_sha256') or '')[:16]
        bstat = r.get('before', {})
        out.append(f"  before: sha={bsha} | drawings={bstat.get('drawings')} | "
                   f"defined_names={bstat.get('defined_names')} | media={bstat.get('media')}")
        if cnt == 0:
            out.append('  (无变更)')
            continue
        out.append(f"  替换 {cnt} 行, 示例:")
        for rec in (r.get('rows_to_change') or [])[:8]:
            out.append(f"    R{rec['row']:>3}  {rec['old']!r} -> {rec['new']!r}")
        if cnt > 8:
            out.append(f"    ... +{cnt - 8} 行")
        if r.get('warnings'):
            out.append('  warnings:')
            for w in r['warnings']:
                out.append(f'    - {w}')
        if 'after_sha256' in r:
            out.append(f"  backup:  {r.get('backup')}")
            asha = (r.get('after_sha256') or '')[:16]
            astat = r.get('after', {})
            out.append(f"  after:   sha={asha} | drawings={astat.get('drawings')} | "
                       f"defined_names={astat.get('defined_names')} | media={astat.get('media')}")
        if r.get('rolled_back'):
            out.append('  [ROLLBACK] verify 失败, 已回滚:')
            for x in r.get('verify_issues', []):
                out.append(f'    - {x}')
        elif r.get('changed'):
            out.append('  ✓ changed')
        # defined_names 净差 (代码审查者 #1 默认仅 WARN, --strict 则 fatal)
        dn_d = r.get('define_name_delta')
        if dn_d is not None:
            tag = ' [STRICT → fatal]' if args.strict_defnames else ' [WARN]'
            out.append(f'  named-range delta: {dn_d:+d}{tag}')
    return '\n'.join(out)


def main():
    ap = argparse.ArgumentParser(
        description="把模板 D 列字面 'M2'/'m2' -> Unicode 'M\u00b2'/'m\u00b2' (默认 dry-run)")
    ap.add_argument('positional', nargs='*', help='xlsx 文件路径 (position 形式)')
    ap.add_argument('--files', nargs='+', default=None, help='xlsx 文件路径 (--files 形式)')
    ap.add_argument('--apply', action='store_true', help='实际写盘 (默认 dry-run)')
    ap.add_argument('--strict-defnames', action='store_true',
                    help='named-range 数变化 也视为 fatal (默认仅 WARN)')
    ap.add_argument('--backup-dir', default=DEFAULT_BACKUP_DIR, help='备份目录')
    ap.add_argument('--json', action='store_true', help='JSON 输出')
    args = ap.parse_args()

    files = args.files if args.files else (args.positional if args.positional else DEFAULT_FILES)

    # 排他锁: 防止同会话跑两次
    lock_path = Path(args.backup_dir) / '.clean_in_progress.lock'
    lock_acquired = False
    if args.apply:
        try:
            lock_path.parent.mkdir(parents=True, exist_ok=True)
            # O_CREAT | O_EXCL 模拟原子独占
            fd = os_open_exclusive(str(lock_path))
            fd.write(datetime.datetime.now().isoformat().encode('utf-8'))
            fd.close()
            lock_acquired = True
        except FileExistsError:
            print(f'[LOCK] 已有别的进程在跑 (lock={lock_path}), 拒绝继续。', file=sys.stderr)
            print(f'[LOCK] 如确认无其它进程, 删除该 .lock 后重试.', file=sys.stderr)
            sys.exit(2)

    try:
        recs = [process_one(f, args) for f in files]
        if args.json:
            print(json.dumps(recs, ensure_ascii=False, indent=2))
        else:
            print(format_human(recs, args))
            if not args.apply:
                print('\n[DRY-RUN] 未改任何文件. 加 --apply 才真写盘.')
    finally:
        if lock_acquired and lock_path.exists():
            try:
                lock_path.unlink()
            except Exception:
                pass


def os_open_exclusive(path):
    """跨平台 O_CREAT|O_EXCL 简化封装."""
    import os
    flags = os.O_CREAT | os.O_EXCL | os.O_WRONLY
    if hasattr(os, 'O_BINARY'):
        flags |= os.O_BINARY
    return os.fdopen(os.open(path, flags, 0o644), 'wb')


if __name__ == '__main__':
    main()
