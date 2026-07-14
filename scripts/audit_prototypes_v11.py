"""v11 audit: 每个 模板 是否都有独立「三 公卫」「四 主卧」「五 主卫」 prototype."""
import sys
import io
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import re

TEMPLATES = ['dizhuan', 'mudiban', 'fushi', 'zhubaojiao']
TEMPLATE_DIR = 'E:\\xiangmu\\baojia\\muban_clean'

# v11 config RoomTypeMaps — must match config.json + QuoteConfig.cs
ROOM_TYPE_MAPS = [
    ("外花园", ["花园", "露台", "庭院", "天井", "花池", "花坛"]),
    ("客餐厅", ["客餐厅", "客厅", "餐厅", "楼梯", "起居", "门厅", "玄关", "过道", "走廊", "入户", "堂屋", "阁楼", "地下室", "吧台", "家庭厅", "阳光房"]),
    ("阳台", ["阳台"]),
    ("主卫", ["主卫", "主卧内卫", "主卧卫生间", "主人卫生间"]),
    ("卫生间", ["卫", "厕所", "洗手间"]),
    ("厨房", ["厨"]),
    ("主卧", ["主卧", "主人房", "主人卧室"]),
    ("卧室", ["卧", "书房", "父母", "儿童", "小孩", "保姆", "衣帽间", "储物间", "茶室", "钢琴房", "老人房", "女儿房", "儿子房", "客房", "电竞房", "影音室", "健身房", "棋牌室", "麻将室", "酒窖", "红酒", "画室", "工作室", "佛堂", "桑拿", "KTV", "多功能室", "休闲厅", "收藏室", "台球室"]),
]

def classify(name):
    for rt, kws in ROOM_TYPE_MAPS:
        if any(kw in name for kw in kws):
            return rt
    return "未分类"

group_re = re.compile(r'^[一二三四五六七八九十]+[、]?$')
KEY_PROTOTYPES = [
    ('三', '公卫', '公卫 bucket 3'),
    ('四', '主卧', '主卧 bucket 4'),
    ('五', '主卫', '主卫 bucket 5'),
]


def audit_template(tpl_path):
    if not os.path.exists(tpl_path):
        return None
    wb = openpyxl.load_workbook(tpl_path, data_only=False)
    ws = wb.active

    prototypes = []
    for row in range(1, ws.max_row + 1):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value
        a_s = str(a).strip() if a is not None else ""
        b_s = str(b).strip() if b is not None else ""
        # 跳过 chrome (八 综合 / 九 其它 / ...)
        if group_re.match(a_s) and b_s and "小计" not in b_s and "合计" not in b_s:
            rtype = classify(b_s)
            prototypes.append((row, a_s, b_s, rtype))
    return prototypes


print("="*72)
print("v11 模板 prototype 结构 审计")
print("="*72)

audit_results = {}

for tpl in TEMPLATES:
    fp = os.path.join(TEMPLATE_DIR, f'{tpl}.xlsx')
    print(f"\n{'='*72}")
    print(f"  模板: {fp}")
    print(f"{'='*72}")
    prototypes = audit_template(fp)
    if prototypes is None:
        print(f"  [ERROR] 文件不存在")
        audit_results[tpl] = None
        continue
    if not prototypes:
        print(f"  [WARN] 没 找到任何 prototype")
        audit_results[tpl] = []
        continue

    print(f"  全部 prototypes ({len(prototypes)} 个):")
    for row, ca, cb, rt in prototypes:
        print(f"    R{row:>3} [{ca:>2} {cb}] → classify={rt}")

    print(f"\n  --- 关键 prototype 审计 ---")
    audit_results[tpl] = []
    for ca_key, cb_key, desc in KEY_PROTOTYPES:
        exact = [p for p in prototypes if p[1] == ca_key and p[2] == cb_key]
        if exact:
            for m in exact:
                print(f"    ✓ 精确: {ca_key} {cb_key} at R{m[0]} → ({desc})")
                audit_results[tpl].append((cb_key, 'exact', m[0]))
        else:
            # check if same col_b exists with different col_a
            alt = [p for p in prototypes if p[2] == cb_key]
            if alt:
                print(f"    ⚠ 变体: {alt[0][1]} {cb_key} at R{alt[0][0]} (不是 {ca_key} {cb_key})")
                audit_results[tpl].append((cb_key, 'variant', alt[0][0]))
            else:
                # check if same col_a exists with different col_b
                alt_a = [p for p in prototypes if p[1] == ca_key]
                if alt_a:
                    print(f"    ⚠ 占位: {ca_key} '{alt_a[0][2]}' at R{alt_a[0][0]} (不是 {cb_key})")
                    audit_results[tpl].append((cb_key, 'occupied', alt_a[0][0]))
                else:
                    print(f"    ✗ 缺失: {ca_key} {cb_key} 完全没找到 → ({desc})")
                    audit_results[tpl].append((cb_key, 'missing', None))

# 汇总表
print()
print("="*72)
print("汇总表")
print("="*72)
print(f"{'模板':<14} | {'三 公卫':<14} | {'四 主卧':<14} | {'五 主卫':<14}")
print("-"*72)
for tpl in TEMPLATES:
    ar = audit_results.get(tpl)
    if ar is None:
        print(f"{tpl:<14} | {'文件缺失':<14} | {'-':<14} | {'-':<14}")
        continue
    cells = []
    for k in ['公卫', '主卧', '主卫']:
        match = next((m for m in ar if m[0] == k), None)
        if not match:
            cells.append('???')
        elif match[1] == 'exact':
            cells.append(f"✓R{match[2]}")
        elif match[1] == 'variant':
            cells.append(f"⚠R{match[2]}(错位)")
        elif match[1] == 'occupied':
            cells.append(f"⚠占位")
        elif match[1] == 'missing':
            cells.append(f"✗缺")
    print(f"{tpl:<14} | {cells[0]:<14} | {cells[1]:<14} | {cells[2]:<14}")

print()
print("="*72)
print("判定: 需要补 五 主卫 prototype 的 模板")
print("="*72)
needs_fix = []
for tpl in TEMPLATES:
    ar = audit_results.get(tpl)
    if ar is None:
        continue
    for k, status, row in ar:
        if k == '主卫' and status != 'exact':
            needs_fix.append(f"  - {tpl}.xlsx: 五 主卫 { '是 变体'+str(row) if status == 'variant' else '占位' if status == 'occupied' else '完全缺失' }")
if needs_fix:
    for line in needs_fix:
        print(line)
else:
    print("  所有 4 模板 都有精确 「五 主卫」 prototype ✓")
