"""UTF-8 safe print of xlsx structure."""
import sys
import os
import io

# Force UTF-8 stdout
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
import re

fp = sys.argv[1] if len(sys.argv) > 1 else None
if not fp or not os.path.exists(fp):
    for c in [
        'C:/Users/Admin（无密码）/Desktop/江南雅苑18#1-2号_报价单_20260714.xlsx',
        '/c/Users/Admin（无密码）/Desktop/江南雅苑18#1-2号_报价单_20260714.xlsx',
    ]:
        if os.path.exists(c):
            fp = c
            break

if not fp:
    import glob
    for g in [
        'C:/Users/*/Desktop/*报价单*.xlsx',
        '/c/Users/*/Desktop/*报价单*.xlsx',
        os.path.expanduser('~') + '/Desktop/*报价单*.xlsx',
    ]:
        m = glob.glob(g)
        if m:
            fp = m[0]
            break

print(f"[OK] {fp}")

wb = openpyxl.load_workbook(fp, data_only=False)
ws = wb.active
print(f"工作表={ws.title} max_row={ws.max_row} max_col={ws.max_column}")
print()

group_a_re = re.compile(r'^[一二三四五六七八九十]+[、]?$')

cur_grp = None
groups = []

for row in range(1, ws.max_row + 1):
    a = ws.cell(row, 1).value
    b = ws.cell(row, 2).value

    # merge check
    merged = False
    for mr in ws.merged_cells.ranges:
        if mr.min_row == row and mr.max_col >= 1 and mr.max_col == ws.max_column:
            # whole row merged
            merged = True
            break

    a_s = str(a).strip() if a is not None else ""
    b_s = str(b).strip() if b is not None else ""

    # floor banner: 整行 merged + col A 含"楼"
    if merged and a_s and "楼" in a_s:
        print(f"[BANNER R{row}] FLOOR=[{a_s}]  full-row merged")
        cur_grp = None

    # 大项
    if group_a_re.match(a_s) and b_s and "小计" not in b_s and "合计" not in b_s:
        cur_grp = (a_s, b_s, row)
        groups.append([cur_grp, [], []])
        print(f"[GROUP R{row}] {a_s} —— {b_s}")

    # 小计 / 合计
    if "小计" in b_s:
        tag = "SUMMARY"
        if cur_grp:
            groups[-1][1].append(row)
            print(f"   [R{row}]    小计 (大项={cur_grp[1]})")
        else:
            print(f"   [R{row}]    小计 (无大项)")

    if "合计" in b_s and "小计" not in b_s:
        if cur_grp:
            groups[-1][2].append(row)
            print(f"   [R{row}]    合计 [{b_s}] (大项={cur_grp[1]})")
        else:
            print(f"   [R{row}]    合计 [{b_s}] (无大项)")

print()
print("=== 大项 + 小计汇总 ===")
for g, subs, totals in groups:
    a, b, r0 = g
    # 检查 col B 在 R{r0+1} 的值 (column 2 of first item row) — this is room name
    row_after_b = ws.cell(r0 + 1, 2).value if r0 + 1 <= ws.max_row else None
    print(f"  R{r0:>3} [{a}] {b}  ↓ 小计行={subs}  ({len(subs)}个)  | first item col2 = '{row_after_b}'")
    # 试着从房间名 判断 RoomType
    if row_after_b:
        rat = ""
        n = str(row_after_b)
        if "主" in n and ("卫" in n or "bath" in n.lower()):
            rat = "(疑似 主卫)"
        elif "公卫" in n or "卫生间" in n or "卫" in n:
            rat = "(公卫)"
        elif "主卧" in n or "主人房" in n:
            rat = "(主卧)"
        elif "卧" in n or "书房" in n or "儿童" in n or "客房" in n:
            rat = "(卧室类)"
        elif "厨房" in n or "厨" in n:
            rat = "(厨房)"
        elif "客" in n or "餐厅" in n:
            rat = "(客餐厅类)"
        elif "阳台" in n:
            rat = "(阳台)"
        elif "花园" in n or "露台" in n:
            rat = "(外花园)"
        if rat:
            print(f"      RoomType guess: {rat}")
