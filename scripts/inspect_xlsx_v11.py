"""
v11 检查: 解析用户生 成的 xlsx, 打印所有大项 header + 小计 / 合计 行.
确认 公卫 / 主卫 / 主卧 三 个 RoomType 是否各占一个 bucket + 各有一个 小计.
"""
import sys
import os
import re
import openpyxl
from collections import defaultdict


def main():
    file_path = sys.argv[1] if len(sys.argv) > 1 else None
    if not file_path:
        print("Usage: inspect_xlsx_v11.py <xlsx_path>")
        print("Trying: E:\\workspace, /tmp, desktop...")
        for cand in [
            'C:/Users/Admin（无密码）/Desktop/江南雅苑18#1-2号_报价单_20260714.xlsx',
            '/c/Users/Admin（无密码）/Desktop/江南雅苑18#1-2号_报价单_20260714.xlsx',
        ]:
            if os.path.exists(cand):
                file_path = cand
                break

    if not file_path or not os.path.exists(file_path):
        # glob search
        import glob
        for pat in [
            'C:/Users/*/Desktop/*报价单*.xlsx',
            '/c/Users/*/Desktop/*报价单*.xlsx',
            '/tmp/*报价单*.xlsx',
            os.path.expanduser('~') + '/Desktop/*报价单*.xlsx',
        ]:
            ms = glob.glob(pat)
            if ms:
                file_path = ms[0]
                break

    if not file_path or not os.path.exists(file_path):
        print(f"ERROR: xlsx 没找到")
        return

    print(f"打开: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb.active
    print(f"工作表: {ws.title}, max_row: {ws.max_row}, max_col: {ws.max_column}")
    print()

    # 大项 header pattern: 列 A = 一/二/三/四/五/六/七/八 + 列 B = 大项名
    group_a_re = re.compile(r'^[一二三四五六七八九十]+[、]?$')

    current_group = None
    groups = []  # [ (a, b, start_row, [subtotal_rows], [total_rows]) ]
    floor_hdr_rows = []  # 复式楼层 banner 行 (整行 merged, e.g. "二楼")

    for row_num in range(1, ws.max_row + 1):
        a = ws.cell(row_num, 1).value
        b = ws.cell(row_num, 2).value
        a_str = str(a).strip() if a is not None else ""
        b_str = str(b).strip() if b is not None else ""

        # 楼层 banner: 列 A merged cell 含 "楼"
        if a_str and ("一楼" in a_str or "二楼" in a_str or "三楼" in a_str or
                      "四楼" in a_str or "五楼" in a_str or "六楼" in a_str):
            print(f"  R{row_num}  FLOOR BANNER: [{a_str}]")

        # 大项 header
        if group_a_re.match(a_str) and b_str and not b_str.startswith("小计") and not b_str.startswith("合计"):
            current_group = (a_str, b_str, row_num)
            groups.append([current_group, [], []])
            print(f"R{row_num}  大项: [{a_str} {b_str}]")

        # 小计
        if "小计" in b_str:
            print(f"  R{row_num}    小计行 → {b_str}  (大项 {'|'.join([g[0][1] for g in groups])[-30:] if groups else '无'})")
            if current_group and groups:
                groups[-1][1].append(row_num)

        # 合计 (非小计)
        if "合计" in b_str and "小计" not in b_str:
            print(f"  R{row_num}    合计行 → {b_str}")
            if current_group and groups:
                groups[-1][2].append(row_num)

    print()
    print("=== 大项 + 小计 汇总 ===")
    for grp_info, sub_rows, total_rows in groups:
        a, b, start = grp_info
        print(f"  [{a} {b}] R{start} → {len(sub_rows)} 个 小计: {sub_rows}")
        if total_rows:
            print(f"      {len(total_rows)} 个 合计: {total_rows}")

    print()
    print("=== 检查 v11 八大类 是否 都出现 ===")
    expected_buckets = ["客餐厅", "厨房", "卫生间", "主卧", "主卫", "卧室", "阳台", "外花园"]
    seen_buckets = {grp[0][1] for grp in groups}
    for eb in expected_buckets:
        present = eb in seen_buckets
        print(f"  [{eb}] {'✓ 出现' if present else '✗ 没出现 (该类 CAD 中无房间)'}")


if __name__ == "__main__":
    main()
