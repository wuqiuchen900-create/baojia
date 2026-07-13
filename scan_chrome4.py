"""
更精细：逐行列出 9 列内容，把 "空白" 准确定位到具体列。
并把整张表"的非空白行 vs 空白行" 映射清楚。
"""
import openpyxl

FILES = [
    r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx",
    r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713-2.xlsx",
]

for fp in FILES:
    print(f"\n{'='*78}\n文件: {fp}\n{'='*78}")
    try:
        wb = openpyxl.load_workbook(fp, data_only=False)
    except Exception as e:
        print(f"加载失败: {e}"); continue

    for ws in wb.worksheets:
        print(f"\n--- Sheet: {ws.title} | max_row={ws.max_row} ---")

        # 找 "八 综合" 起始行
        chrome_start = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if v and "八 " in str(v) and "综合" in str(v):
                chrome_start = r
                break
        # 找 "九 " or 九 其它 段
        after_chrome_start = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 2).value
            if v and (str(v).startswith("九 ") or str(v).strip() == "其它"):
                after_chrome_start = r
                break

        print(f"  八综合 起始 = {chrome_start}, 九 段 起始 = {after_chrome_start}")

        if chrome_start is None or after_chrome_start is None:
            continue

        span_start = chrome_start
        span_end = min(after_chrome_start + 12, ws.max_row)

        print(f"\n  R{span_start}..R{span_end} 逐行 9 列:")
        print(f"  {'R':>4} | c1 | c2 (name) | c3 | c4 | c5 | c6 | c7 | c8 | c9")
        print(f"  --+--+--+--+--+--+--+--+--+--")

        for r in range(span_start, span_end + 1):
            contents = [ws.cell(r, c).value for c in range(1, 10)]
            label_cells = []
            for v in contents:
                if v in (None, ""):
                    label_cells.append(".")
                else:
                    s = str(v).strip()[:18]
                    label_cells.append(s)
            print(f"  {r:>4} | {' | '.join(label_cells)}")

        # 计算 R(chrome_start)..R(after_chrome_start) 范围内空白行分布
        print(f"\n  R{span_start}..R{after_chrome_start} 范围内按 9 列分组的空白情况:")
        empty_2to6 = []
        empty_3to8 = []
        empty_almost_all = []
        cur_2to6 = None; cur_2to6_s = None
        cur_3to8 = None; cur_3to8_s = None
        cur_all = None; cur_all_s = None
        for r in range(span_start, after_chrome_start + 1):
            vs = [ws.cell(r, c).value for c in range(1, 10)]
            c2to6 = [vs[i] for i in range(1, 6)]  # col2..col6
            c3to8 = [vs[i] for i in range(2, 8)]  # col3..col8
            c2to6_blank = all(v in (None,"") for v in c2to6)
            c3to8_blank = all(v in (None,"") for v in c3to8)
            almost_all = all(v in (None,"") for v in vs[1:])  # col2..col9 all blank (col1 ignored)
            if c2to6_blank:
                if cur_2to6 is None: cur_2to6_s = r
                cur_2to6 = r
            else:
                if cur_2to6 is not None:
                    empty_2to6.append((cur_2to6_s, cur_2to6)); cur_2to6 = None
            if c3to8_blank:
                if cur_3to8 is None: cur_3to8_s = r
                cur_3to8 = r
            else:
                if cur_3to8 is not None:
                    empty_3to8.append((cur_3to8_s, cur_3to8)); cur_3to8 = None
            if almost_all:
                if cur_all is None: cur_all_s = r
                cur_all = r
            else:
                if cur_all is not None:
                    empty_almost_all.append((cur_all_s, cur_all)); cur_all = None
        if cur_2to6 is not None: empty_2to6.append((cur_2to6_s, cur_2to6))
        if cur_3to8 is not None: empty_3to8.append((cur_3to8_s, cur_3to8))
        if cur_all  is not None: empty_almost_all.append((cur_all_s, cur_all))

        print(f"  blank col2..col6 区间: {empty_2to6}")
        print(f"  blank col3..col8 区间: {empty_3to8}")
        print(f"  blank col2..col9 区间 (col2..9 全空): {empty_almost_all}")
