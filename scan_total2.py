"""完整 dump 整张表:列出所有 col2 含"小计"的行, AND 整张表 9 列内容中所有 col2 文本非空的行 (T=12..160)"""
import openpyxl

fp = r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx"
wb = openpyxl.load_workbook(fp, data_only=False)

for ws in wb.worksheets:
    print(f"\n==== Sheet: {ws.title} | max_row={ws.max_row} ====")

    # 第一部分: 列出 ALL col2 含 "小计" OR "合计" 的行
    print("\n--- A. 所有 col2 含 '小计' 或 '合计' ---")
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v is None: continue
        sv = str(v).strip()
        if "小计" in sv or "合计" in sv:
            c1 = ws.cell(r, 1).value
            cf = ws.cell(r, 6).value
            ch = ws.cell(r, 8).value
            print(f"  R{r:>4} | col1='{str(c1).strip() if c1 else '.'}' | col2='{sv}' | F={cf} | H={ch}")

    # 第二部分: 行 R12..R160 全 9 列 dump
    print(f"\n--- B. R12..R160 完整 9 列 ---")
    for r in range(12, min(161, ws.max_row + 1)):
        parts = []
        for c in range(1, 10):
            v = ws.cell(r, c).value
            if v is None: parts.append(".")
            else: parts.append(str(v).strip()[:18])
        print(f"  R{r:>4} | {' | '.join(parts)}")
