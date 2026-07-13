"""逐行列 9 列内容到 chrome_dump.txt 文件（避免 basher summarization 吃字段）"""
import openpyxl

FILES = [
    r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx",
    r"C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713-2.xlsx",
]

OUT = r"E:\xiangmu\baojia\chrome_dump.txt"

with open(OUT, "w", encoding="utf-8") as out:
    for fp in FILES:
        out.write(f"\n{'='*80}\n文件: {fp}\n{'='*80}\n")
        wb = openpyxl.load_workbook(fp, data_only=False)
        for ws in wb.worksheets:
            out.write(f"\n--- Sheet: {ws.title} | max_row={ws.max_row} max_col={ws.max_column} ---\n")

            # 找 chrome 起点
            chrome_start = None
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, 2).value
                if v and "八 " in str(v) and "综合" in str(v):
                    chrome_start = r; break

            # 找 chrome 终点 (找 "九 其它" 或 "九 " 任意出现)
            chrome_end = None
            for r in range((chrome_start or 1), ws.max_row + 1):
                v = ws.cell(r, 2).value
                if v and "九 " in str(v):
                    chrome_end = r; break

            out.write(f"chrome 八综合 起始 = {chrome_start},  九 起始 = {chrome_end}\n")
            if chrome_start is None or chrome_end is None:
                continue

            # 整个 chrome 区 + 之后 3 行
            span_s = chrome_start
            span_e = min(chrome_end + 5, ws.max_row)
            out.write(f"\n逐行 R{span_s}..R{span_e} (9 列):\n")
            out.write(f"{'R':>4} | c1 | c2                  | c3     | c4     | c5     | c6            | c7        | c8       | c9\n")
            out.write(f"{'-'*4}-+-{'-'*2}-+-{'-'*19}-+-{'-'*6}-+-{'-'*6}-+-{'-'*6}-+-{'-'*13}-+-{'-'*9}-+-{'-'*8}-+-{'-'*6}\n")
            for r in range(span_s, span_e + 1):
                cells = [ws.cell(r, c).value for c in range(1, 10)]
                parts = []
                for v in cells:
                    if v in (None, ""):
                        parts.append(".")
                    else:
                        parts.append(str(v).strip()[:max(2, 18 if cells.index(v)==1 else 8)])
                out.write(f"{r:>4} | {' | '.join(parts)}\n")

            # 区间内空白归类
            out.write(f"\n区间 R{span_s}..R{chrome_end} 空白归类:\n")
            runs_2to6 = []; runs_3to8 = []; runs_almost = []
            s26 = e26 = None; s38 = e38 = None; sa = ea = None
            for r in range(span_s, chrome_end + 1):
                vs = [ws.cell(r, c).value for c in range(1, 10)]
                v26 = [vs[i] for i in range(1, 6)]
                v38 = [vs[i] for i in range(2, 8)]
                v_ = vs[1:]  # col2..col9
                b26 = all(v in (None,"") for v in v26)
                b38 = all(v in (None,"") for v in v38)
                ba  = all(v in (None,"") for v in v_)
                if b26:
                    if s26 is None: s26 = r
                    e26 = r
                elif s26 is not None:
                    runs_2to6.append((s26, e26)); s26 = e26 = None
                if b38:
                    if s38 is None: s38 = r
                    e38 = r
                elif s38 is not None:
                    runs_3to8.append((s38, e38)); s38 = e38 = None
                if ba:
                    if sa is None: sa = r
                    ea = r
                elif sa is not None:
                    runs_almost.append((sa, ea)); sa = ea = None
            if s26 is not None: runs_2to6.append((s26, e26))
            if s38 is not None: runs_3to8.append((s38, e38))
            if sa  is not None: runs_almost.append((sa, ea))
            out.write(f"  blank col2..col6:        {runs_2to6}\n")
            out.write(f"  blank col3..col8:        {runs_3to8}\n")
            out.write(f"  blank col2..col9 (忽略col1): {runs_almost}\n")

# Also write a summary at the end
print(f"已写入 {OUT}")
import os
print(f"大小: {os.path.getsize(OUT)} bytes")
