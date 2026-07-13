# -*- coding: utf-8 -*-
import openpyxl, sys, zipfile, re
sys.stdout.reconfigure(encoding='utf-8')

path = r'C:\Users\Admin（无密码）\Desktop\江南雅苑18#1-2号_报价单_20260713.xlsx'
print(f"=== 扫描: {path} ===\n")

# 1. xlsx 包内看 logo / drawings / media
print("=== [Logo / Drawings / Media] ===")
with zipfile.ZipFile(path) as z:
    media = sorted([n for n in z.namelist() if 'media' in n.lower()])
    drawings = sorted([n for n in z.namelist() if 'drawing' in n.lower()])
    print(f"  media files: {media}")
    print(f"  drawing files: {drawings}")
print()

# 2. Open sheet
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb.active
print(f"=== 工作表: {ws.title}, 总行数: {ws.max_row} ===\n")

# 3. 列出所有 "小计" / "合计" / "楼" / "综合" / "其它" 行 + 公式
print("=== [关键行 + 公式] ===")
he_keywords = ('小计', '合计', '直接费', '综合', '其它', '主材', '楼')
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9, values_only=False):
    r = row[0].row
    c1 = str(row[0].value or '').strip()
    c2 = str(row[1].value or '').strip()
    if any(kw in c2 for kw in he_keywords) or c1 in ('一','二','三','四','五','六','七','八','九','十'):
        # 检查 6/8 列 公式
        f6 = str(ws.cell(row=r, column=6).value or '').strip()
        f8 = str(ws.cell(row=r, column=8).value or '').strip()
        marker = ''
        if f6 and f6.startswith('='): marker += f'  F=[{f6[:50]}]'
        if f8 and f8.startswith('='): marker += f'  H=[{f8[:50]}]'
        # 检查 3 列 (数量)
        c3 = ws.cell(row=r, column=3).value
        if c3 is not None and not isinstance(c3, str):
            marker += f'  C={c3}'
        elif c3:
            marker += f'  C={c3[:20]}'
        print(f"R{r:3d} | {c1[:4]:4s} | {c2[:30]:30s}{marker}")
print()

# 4. 数 chrome "其它" 段的 item 数量 + 公式分布
print("=== [chrome 其它 段统计] ===")
other_start = None
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2, values_only=False):
    r = row[0].row
    c1 = str(row[0].value or '').strip()
    c2 = str(row[1].value or '').strip()
    if c1 == '九' and c2.startswith('其它'):
        other_start = r
        print(f"  九其它 标题行: R{r}")
        break
if other_start:
    item_count = 0
    formula_count = 0
    for r in range(other_start+1, ws.max_row+1):
        c1 = str(ws.cell(row=r, column=1).value or '').strip()
        c2 = str(ws.cell(row=r, column=2).value or '').strip()
        f6 = str(ws.cell(row=r, column=6).value or '').strip()
        if c1.isdigit() or c2:
            item_count += 1
            if f6.startswith('='):
                formula_count += 1
        if '合计' in c2 or c1 in ('十','十一'):
            break
    print(f"  九其它 段 item 数: {item_count}, 含公式数: {formula_count}")

# 5. 数所有 col6/8 是公式的行数 与 小计行数
print("\n=== [全表 col6/8 公式统计] ===")
f6cnt = f8cnt = 0
subtotal_rows = []
total_rows = []
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, max_col=9, values_only=False):
    r = row[0].row
    c2 = str(ws.cell(row=r, column=2).value or '').strip()
    f6 = str(ws.cell(row=r, column=6).value or '').strip()
    f8 = str(ws.cell(row=r, column=8).value or '').strip()
    if f6.startswith('='): f6cnt += 1
    if f8.startswith('='): f8cnt += 1
    if '小计' in c2: subtotal_rows.append(r)
    if '合计' in c2 and '小计' not in c2: total_rows.append(r)
print(f"  col6 公式行数: {f6cnt}, col8 公式行数: {f8cnt}")
print(f"  小计行 ({len(subtotal_rows)} 个): {subtotal_rows}")
print(f"  合计行 ({len(total_rows)} 个): {total_rows}")

# 6. 最后一行的合计公式内容
print("\n=== [最后一行 (合计) 公式详情] ===")
if total_rows:
    last_total = max(total_rows)
    for col_idx, col_name in [(1,'A'),(2,'B'),(3,'C'),(4,'D'),(5,'E'),(6,'F'),(7,'G'),(8,'H'),(9,'I')]:
        v = ws.cell(row=last_total, column=col_idx).value
        print(f"  R{last_total} [{col_name}]: {v}")
